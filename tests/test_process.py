import pandas as pd
import pytest
from process import calculate_student_goals


def make_df(rows):
    """rows: list of (student_name, pages, date_str, center)"""
    return pd.DataFrame(rows, columns=["Student Name", "Pages Completed", "Date", "Center"])


# ── Standard student (10+ sessions) ──────────────────────────────────────────

def test_standard_student_uses_last_10():
    rows = [("Alice", i, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 16)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["num_sessions"] == 10
    assert len([s for s in student["sessions"] if s is not None]) == 10


def test_standard_student_uses_most_recent_10():
    rows = [("Alice", i, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 16)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    # Most recent 10 sessions: pages 6–15, oldest first
    non_null = [s for s in student["sessions"] if s is not None]
    assert non_null[0] == 6   # oldest of the 10
    assert non_null[-1] == 15  # most recent


def test_standard_student_average():
    # 10 sessions with pages 1-10
    rows = [("Bob", i, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 11)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["average"] == round(sum(range(1, 11)) / 10, 2)  # 5.5


def test_standard_student_goal_is_120_pct_of_average():
    rows = [("Bob", i, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 11)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    expected_goal = round(5.5 * 1.20, 2)  # 6.6
    assert student["goal"] == expected_goal


# ── High performer: goal capped below max ─────────────────────────────────────

def test_goal_capped_below_max():
    # 10 sessions all with value 10 → avg=10, 120%=12, max=10, goal=9.99
    rows = [("Carol", 10, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 11)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["goal"] == 9.99


def test_goal_uncapped_when_120pct_is_below_max():
    # sessions: 5,5,5,5,5,10,10,10,10,10 → avg=7.5, 120%=9.0, max=10, goal=9.0
    # (120% of avg = 9.0 < max-0.01 = 9.99, so goal = 9.0, no cap applied)
    pages = [5] * 5 + [10] * 5
    rows = [("Dave", p, f"2026-03-{i:02d}", "Teaneck") for i, p in enumerate(pages, 1)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["goal"] == 9.0


# ── Zeros excluded ────────────────────────────────────────────────────────────

def test_zeros_excluded():
    rows = [
        ("Eve", 0, "2026-03-01", "Teaneck"),
        ("Eve", 0, "2026-03-02", "Teaneck"),
        ("Eve", 8, "2026-03-03", "Teaneck"),
        ("Eve", 9, "2026-03-04", "Teaneck"),
    ]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["num_sessions"] == 2
    non_null = [s for s in student["sessions"] if s is not None]
    assert 0 not in non_null


def test_student_with_all_zeros_excluded():
    rows = [("Frank", 0, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 6)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    assert len(results["Teaneck"]) == 0


# ── Null pages excluded ────────────────────────────────────────────────────────

def test_null_pages_excluded():
    rows = [
        ("Grace", None, "2026-03-01", "Teaneck"),
        ("Grace", 7,    "2026-03-02", "Teaneck"),
        ("Grace", 8,    "2026-03-03", "Teaneck"),
    ]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["num_sessions"] == 2


# ── Partial data (fewer than 10 sessions) ─────────────────────────────────────

def test_partial_student_uses_all_sessions():
    rows = [("Hal", i, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 4)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert student["num_sessions"] == 3
    non_null = [s for s in student["sessions"] if s is not None]
    assert len(non_null) == 3


def test_partial_student_sessions_padded_with_none():
    rows = [("Hal", i, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 4)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    student = results["Teaneck"][0]
    assert len(student["sessions"]) == 10
    assert student["sessions"].count(None) == 7


# ── Center assignment ─────────────────────────────────────────────────────────

def test_englewood_center_assignment():
    rows = [("Ivy", 8, f"2026-03-{i:02d}", "Englewood") for i in range(1, 5)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    assert len(results["Englewood"]) == 1
    assert len(results["Teaneck"]) == 0


def test_teaneck_virtual_assigned_to_teaneck():
    rows = [("Jay", 8, f"2026-03-{i:02d}", "Teaneck, Teaneck Virtual") for i in range(1, 5)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    assert len(results["Teaneck"]) == 1


def test_englewood_virtual_assigned_to_englewood():
    rows = [("Kim", 8, f"2026-03-{i:02d}", "Englewood, Teaneck Virtual") for i in range(1, 5)]
    df = make_df(rows)
    results = calculate_student_goals(df)
    assert len(results["Englewood"]) == 1


# ── Alphabetical sort ─────────────────────────────────────────────────────────

def test_students_sorted_alphabetically():
    rows = (
        [("Zara", 8, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 5)] +
        [("Aaron", 8, f"2026-03-{i:02d}", "Teaneck") for i in range(1, 5)]
    )
    df = make_df(rows)
    results = calculate_student_goals(df)
    names = [s["name"] for s in results["Teaneck"]]
    assert names == sorted(names)


def test_write_excel_creates_file_with_correct_sheets(tmp_path):
    import openpyxl
    results = {
        "Englewood": [
            {"name": "Alice", "num_sessions": 2, "sessions": [None]*8 + [5.0, 6.0], "average": 5.5, "goal": 5.49}
        ],
        "Teaneck": [],
    }
    out = str(tmp_path / "test_output.xlsx")
    from process import write_excel
    write_excel(results, out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Englewood", "Teaneck"]
    ws = wb["Englewood"]
    assert ws.cell(1, 1).value == "Student Name"
    assert ws.cell(2, 1).value == "Alice"
    assert ws.cell(1, 14).value == "Page Goal"
