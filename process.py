import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def calculate_student_goals(df: pd.DataFrame) -> dict[str, list[dict]]:
    """
    Given the full DWP DataFrame, return a dict:
      {"Englewood": [student_record, ...], "Teaneck": [student_record, ...]}

    Each student_record:
      {
        "name": str,
        "num_sessions": int,
        "sessions": list[float | None],  # always length 10, oldest first
        "average": float,
        "goal": float,
      }
    """
    df = df[["Student Name", "Pages Completed", "Date", "Center"]].copy()
    df["Date"] = pd.to_datetime(df["Date"])

    # Filter: Pages Completed must be numeric and > 0
    df = df[pd.to_numeric(df["Pages Completed"], errors="coerce").notna()]
    df["Pages Completed"] = df["Pages Completed"].astype(float)
    df = df[df["Pages Completed"] > 0]

    # Assign center
    def assign_center(center_str: str) -> str | None:
        # Englewood is checked first — "Englewood, Teaneck Virtual" routes to Englewood
        if "Englewood" in str(center_str):
            return "Englewood"
        if "Teaneck" in str(center_str):
            return "Teaneck"
        return None

    df["AssignedCenter"] = df["Center"].apply(assign_center)
    df = df[df["AssignedCenter"].notna()]

    results = {"Englewood": [], "Teaneck": []}

    for (name, center), group in df.groupby(["Student Name", "AssignedCenter"]):
        # Sort by date descending, take most recent 10
        group = group.sort_values("Date", ascending=False).head(10)
        pages = group["Pages Completed"].tolist()
        num_sessions = len(pages)

        # Reverse so oldest is first (Session 1 = oldest)
        pages_oldest_first = list(reversed(pages))

        average = round(sum(pages_oldest_first) / num_sessions, 2)
        max_pages = max(pages_oldest_first)
        # Goal is 20% above average, capped just below the student's personal best
        # (max - 0.01 keeps the goal achievable — always something to reach for)
        goal = round(min(average * 1.20, max_pages - 0.01), 2)

        # Pad to 10 with None (left-pad: Nones go in early session slots)
        padded = [None] * (10 - num_sessions) + pages_oldest_first

        results[center].append({
            "name": name,
            "num_sessions": num_sessions,
            "sessions": padded,
            "average": average,
            "goal": goal,
        })

    # Sort alphabetically
    for center in results:
        results[center].sort(key=lambda s: s["name"])

    return results


def write_excel(results: dict[str, list[dict]], output_path: str, raw_df: pd.DataFrame = None) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill   = PatternFill("solid", fgColor="4472C4")
    session_fill  = PatternFill("solid", fgColor="E7E6E6")
    summary_fill  = PatternFill("solid", fgColor="FFF2CC")
    header_font   = Font(bold=True, color="FFFFFF")
    center_align  = Alignment(horizontal="center")
    left_align    = Alignment(horizontal="left")

    headers = (
        ["Student Name", "# Sessions"] +
        [f"Session {i}" for i in range(1, 11)] +
        ["Average Pages", "Page Goal"]
    )
    col_widths = [30, 18] + [10] * 10 + [14, 12]

    for sheet_name in ["Englewood", "Teaneck"]:
        ws = wb.create_sheet(sheet_name)
        students = results[sheet_name]

        # Header row
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, student in enumerate(students, start=2):
            ws.cell(row=row_idx, column=1, value=student["name"]).alignment = left_align
            ws.cell(row=row_idx, column=2, value=student["num_sessions"]).alignment = center_align

            for s_idx, pages in enumerate(student["sessions"]):
                cell = ws.cell(row=row_idx, column=3 + s_idx, value=pages)
                cell.fill = session_fill
                cell.alignment = center_align

            avg_cell = ws.cell(row=row_idx, column=13, value=student["average"])
            avg_cell.fill = summary_fill
            avg_cell.alignment = center_align

            goal_cell = ws.cell(row=row_idx, column=14, value=student["goal"])
            goal_cell.fill = summary_fill
            goal_cell.alignment = center_align

        ws.freeze_panes = "B2"

    # Raw data tab
    if raw_df is not None:
        ws = wb.create_sheet("Data")
        # Header row
        for col_idx, col_name in enumerate(raw_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        # Data rows
        for row_idx, row in enumerate(raw_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        ws.freeze_panes = "A2"

    wb.save(output_path)


def process_report(input_path: str, output_path: str) -> None:
    df = pd.read_excel(input_path)
    results = calculate_student_goals(df)
    write_excel(results, output_path, raw_df=df)
